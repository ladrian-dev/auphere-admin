"""Importa un catálogo de productos a ``local_catalog_products``.

Lee un CSV, un JSON o un XLSX y deja las filas listas para que el backend
``catalogo_local`` de las tools ``inventory.*`` las sirva. Es el puente que
permite tener el agente funcionando mientras el POS de origen no tiene datos.

**Los nombres de columna se reconocen por alias**, porque el fichero de un
negocio real nunca viene con los nuestros: ``producto``/``descripcion`` valen
como ``nombre``, ``codigo`` como ``sku``, ``existencia`` como ``stock``, etc.
La comparación ignora mayúsculas, tildes y espacios.

**El precio de costo se descarta si viene.** La tabla no tiene esa columna a
propósito (ver la migración 0099).

``--file`` acepta una ruta local o un URI ``s3://bucket/clave``. Lo segundo
existe para los entornos desplegados: la base vive en subred privada, así que
esto se corre como tarea ECS puntual (igual que ``nexus-migrate``) y ahí no
hay ficheros locales que montar.

Uso:

    python apps/api/scripts/load_local_catalog.py \\
        --tenant-slug demo-farmacia-amigable \\
        --file productos.csv --replace

    # en un entorno desplegado
    python apps/api/scripts/load_local_catalog.py \\
        --tenant-slug demo-farmacia-amigable \\
        --file s3://nexus-prod-media/imports/catalogo.csv --replace
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import async_sessionmaker

from nexus_api.core.tenant_context import tenant_scoped_session
from nexus_api.db.base import get_engine
from nexus_api.db.models.tenant import Tenant

# El plegado se IMPORTA del cliente que luego lee estas filas, no se
# reimplementa: si el importador y el lector pliegan distinto, las filas
# quedan escritas pero inbuscables, y el agente responde "no lo tengo"
# sobre un producto que sí está.
from nexus_mcp.servers.catalogo_local.client import fold


# Alias -> columna canónica. La clave se compara ya plegada.
_ALIASES: dict[str, str] = {
    "nombre": "nombre", "producto": "nombre", "descripcion": "nombre",
    "descripción": "nombre", "articulo": "nombre", "item": "nombre",
    "sku": "sku", "codigo": "sku", "código": "sku", "cod": "sku",
    "referencia": "sku", "ref": "sku",
    "categoria": "categoria", "categoría": "categoria", "familia": "categoria",
    "linea": "categoria", "grupo": "categoria",
    "tipo": "tipo",
    "precio_usd": "precio_usd", "precio": "precio_usd", "pvp": "precio_usd",
    "precio venta": "precio_usd", "precio_venta": "precio_usd", "usd": "precio_usd",
    "stock_actual": "stock_actual", "stock": "stock_actual",
    "existencia": "stock_actual", "existencias": "stock_actual",
    "cantidad": "stock_actual", "disponible": "stock_actual",
    "stock_minimo": "stock_minimo", "minimo": "stock_minimo",
    "mínimo": "stock_minimo", "stock min": "stock_minimo",
    "punto_reorden": "stock_minimo",
}


def _canonical(header: str) -> str | None:
    return _ALIASES.get(fold(header).replace("_", " ").replace("  ", " ")) or _ALIASES.get(fold(header))


def _to_decimal(raw: Any) -> Decimal:
    s = str(raw or "0").strip().replace("$", "").replace(" ", "")
    if "," in s and "." in s:          # 1.234,56 -> 1234.56
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:                      # 12,50 -> 12.50
        s = s.replace(",", ".")
    try:
        return Decimal(s or "0").quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")


def _to_int(raw: Any) -> int:
    s = str(raw or "0").strip().replace(".", "").replace(",", "")
    try:
        return int(float(s or 0))
    except ValueError:
        return 0


def _materialise(source: str) -> Path:
    """Devuelve una ruta local, descargando de S3 si hace falta."""
    if not source.startswith("s3://"):
        return Path(source)
    import tempfile

    import boto3

    bucket, _, key = source[len("s3://"):].partition("/")
    if not bucket or not key:
        print(f"ERROR: URI de S3 mal formado: {source}")
        raise SystemExit(2)
    suffix = Path(key).suffix or ".csv"
    tmp = Path(tempfile.mkstemp(suffix=suffix)[1])
    boto3.client("s3").download_file(bucket, key, str(tmp))
    print(f"descargado {source} ({tmp.stat().st_size} bytes)")
    return tmp


def _read_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        rows = payload.get("data") if isinstance(payload, dict) else payload
        return [r for r in (rows or []) if isinstance(r, dict)]
    if suffix in (".csv", ".tsv", ".txt"):
        delim = "\t" if suffix == ".tsv" else ","
        with path.open(encoding="utf-8-sig", newline="") as fh:
            return list(csv.DictReader(fh, delimiter=delim))
    if suffix in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            print(
                "ERROR: falta openpyxl para leer XLSX. "
                "Exporta la hoja a CSV y vuelve a intentarlo."
            )
            raise SystemExit(2) from None
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        headers = [str(h or "") for h in next(it)]
        return [dict(zip(headers, r, strict=False)) for r in it]
    print(f"ERROR: extensión no soportada: {suffix} (usa .csv, .json o .xlsx)")
    raise SystemExit(2)


def _normalise(raw_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    """Devuelve (filas normalizadas, avisos)."""
    warnings: list[str] = []
    if not raw_rows:
        return [], ["el fichero no tiene filas"]

    mapping = {h: _canonical(h) for h in raw_rows[0]}
    known = {v for v in mapping.values() if v}
    ignored = sorted(h for h, c in mapping.items() if c is None and h)
    if ignored:
        warnings.append(f"columnas ignoradas: {', '.join(ignored)}")
    for required in ("nombre", "sku"):
        if required not in known:
            warnings.append(f"FALTA la columna obligatoria {required!r}")

    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for raw in raw_rows:
        row: dict[str, Any] = {}
        for header, canon in mapping.items():
            if canon:
                row[canon] = raw.get(header)
        nombre = str(row.get("nombre") or "").strip()
        sku = str(row.get("sku") or "").strip()
        if not nombre or not sku:
            continue
        if sku in seen:
            warnings.append(f"SKU duplicado en el fichero, se queda el primero: {sku}")
            continue
        seen.add(sku)
        out.append(
            {
                "sku": sku,
                "nombre": nombre,
                "categoria": (str(row.get("categoria") or "").strip() or None),
                "tipo": (str(row.get("tipo") or "").strip() or None),
                "precio_usd": _to_decimal(row.get("precio_usd")),
                "stock_actual": _to_int(row.get("stock_actual")),
                "stock_minimo": _to_int(row.get("stock_minimo")),
                "search_text": fold(f"{nombre} {sku}"),
            }
        )
    return out, warnings


async def _amain(args: argparse.Namespace) -> int:
    path = _materialise(args.file)
    if not path.exists():
        print(f"ERROR: no existe {path}")
        return 1

    rows, warnings = _normalise(_read_rows(path))
    for w in warnings:
        print(f"aviso: {w}")
    if not rows:
        print("ERROR: no se pudo normalizar ninguna fila (¿faltan 'nombre' y 'sku'?)")
        return 1

    engine = get_engine()
    Session = async_sessionmaker(engine, expire_on_commit=False)

    async with Session() as lookup:
        tenant = (
            await lookup.execute(select(Tenant).where(Tenant.slug == args.tenant_slug))
        ).scalar_one_or_none()
    if tenant is None:
        print(f"ERROR: tenant {args.tenant_slug!r} no encontrado")
        return 1

    async with Session() as session, tenant_scoped_session(session, tenant.id):
        if args.replace:
            deleted = await session.execute(
                text("DELETE FROM local_catalog_products WHERE tenant_id = :t"),
                {"t": str(tenant.id)},
            )
            print(f"borradas {deleted.rowcount} filas previas (--replace)")

        for row in rows:
            await session.execute(
                text(
                    """
                    INSERT INTO local_catalog_products (
                        tenant_id, sku, nombre, categoria, tipo,
                        precio_usd, stock_actual, stock_minimo, search_text
                    ) VALUES (
                        :tenant_id, :sku, :nombre, :categoria, :tipo,
                        :precio_usd, :stock_actual, :stock_minimo, :search_text
                    )
                    ON CONFLICT (tenant_id, sku) DO UPDATE SET
                        nombre = EXCLUDED.nombre,
                        categoria = EXCLUDED.categoria,
                        tipo = EXCLUDED.tipo,
                        precio_usd = EXCLUDED.precio_usd,
                        stock_actual = EXCLUDED.stock_actual,
                        stock_minimo = EXCLUDED.stock_minimo,
                        search_text = EXCLUDED.search_text,
                        updated_at = now()
                    """
                ),
                {"tenant_id": str(tenant.id), **row},
            )

        total = (
            await session.execute(
                text("SELECT count(*) FROM local_catalog_products WHERE tenant_id = :t"),
                {"t": str(tenant.id)},
            )
        ).scalar_one()
        bajo = (
            await session.execute(
                text(
                    "SELECT count(*) FROM local_catalog_products "
                    "WHERE tenant_id = :t AND stock_actual <= stock_minimo"
                ),
                {"t": str(tenant.id)},
            )
        ).scalar_one()

    print(f"{args.tenant_slug}: {len(rows)} filas importadas · {total} en catálogo · {bajo} bajo mínimo")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="Importa un catálogo a local_catalog_products")
    ap.add_argument("--tenant-slug", required=True)
    ap.add_argument(
        "--file",
        required=True,
        help="CSV, JSON o XLSX — ruta local o s3://bucket/clave",
    )
    ap.add_argument(
        "--replace",
        action="store_true",
        help="borra el catálogo previo del tenant antes de importar",
    )
    return asyncio.run(_amain(ap.parse_args()))


if __name__ == "__main__":
    raise SystemExit(main())
